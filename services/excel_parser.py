"""
Vertex42 캘린더 엑셀 템플릿 파서
"""
import re
from datetime import datetime, timedelta
import openpyxl
import logging

logger = logging.getLogger(__name__)

# 건너뛸 시트 (정보 페이지)
INFO_SHEETS = {'정보', '정보 (2)', 'Sheet1'}

# 주 시작 행 (6행 간격)
WEEK_START_ROWS = [10, 16, 22, 28, 34, 40]

# 요일 설정: (날짜 열 인덱스, 수업명 열 인덱스)
# Mon=C(3)/D(4), Tue=E(5)/F(6), Wed=G(7)/H(8), Thu=I(9)/J(10), Fri=K(11)/M(13)
DAY_CONFIG = [
    (3, 4),    # Monday:   C=날짜, D=수업명
    (5, 6),    # Tuesday:  E=날짜, F=수업명
    (7, 8),    # Wednesday: G=날짜, H=수업명
    (9, 10),   # Thursday: I=날짜, J=수업명
    (11, 13),  # Friday:   K=날짜, M=수업명
]

# 공휴일/비수업 키워드
HOLIDAY_KEYWORDS = {
    '추석', '개천절', '한글날', '대체휴일', '방학', '어린이날',
    '현충일', '광복절', '석가탄신일', '삼일절', '신정', '성탄절',
    '구정', '설날', '연휴', '휴일', '메모', '새해',
}


def get_sheet_names(filepath):
    """엑셀 파일의 시트 목록 반환 (정보 시트 제외)"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheets = [name for name in wb.sheetnames if name not in INFO_SHEETS]
    wb.close()
    return sheets


def calculate_end_time(start_time_str, hours):
    """시작 시간 + 수업 시간 → 종료 시간 계산 (점심시간 13:00~14:00 포함)"""
    h, m = map(int, start_time_str.split(':'))
    start = datetime(2000, 1, 1, h, m)
    end = start + timedelta(hours=hours)

    # 점심시간 (13:00~14:00) 걸치면 1시간 추가
    lunch_start = datetime(2000, 1, 1, 13, 0)
    if start < lunch_start and end > lunch_start:
        end += timedelta(hours=1)

    return end.strftime('%H:%M')


def _detect_week_start_rows(ws):
    """시트에서 실제 주 시작 행을 동적으로 감지"""
    date_rows = set()
    for row in range(8, 55):
        for date_col, _ in DAY_CONFIG:
            cell_val = ws.cell(row=row, column=date_col).value
            if isinstance(cell_val, datetime) and cell_val.year >= 2020:
                date_rows.add(row)
                break

    if not date_rows:
        logger.warning("날짜 행을 감지하지 못했습니다. 기본값 사용.")
        return WEEK_START_ROWS  # fallback

    # 인접 행 그룹핑: 4행 이상 떨어져 있으면 새로운 주
    sorted_rows = sorted(date_rows)
    week_rows = [sorted_rows[0]]
    for r in sorted_rows[1:]:
        if r - week_rows[-1] >= 4:
            week_rows.append(r)

    logger.info(f"감지된 주 시작 행: {week_rows}")
    return week_rows


NOT_NAME_WORDS = {
    '발표', '자격증', '주제별', '네트워킹', '온라인', '오프라인',
    '자기소개', '팀구성', '특강', '보강', '실습', '복습',
    '평가', '시험', '면접', '상담', '수료', '졸업',
    '쇼츠', '영상', '촬영', '편집', '기획', '운영',
}


def _is_korean_name(text):
    """한국어 강사 이름인지 판별 (2~4자 한글, 설명 단어 제외)"""
    if not text or not isinstance(text, str):
        return False
    cleaned = text.strip()
    if not cleaned:
        return False
    # 숫자만 있는 경우 제외
    if cleaned.replace('h', '').replace('H', '').replace(' ', '').isdigit():
        return False
    # '8h' 같은 시간 형식 제외
    if re.match(r'^\d+h?$', cleaned, re.IGNORECASE):
        return False
    # 이름이 아닌 키워드 필터
    if cleaned in NOT_NAME_WORDS:
        return False
    # 한국어 이름: 2~4자 한글만 (한국 이름은 대부분 2~4자)
    if re.match(r'^[가-힣]{2,4}$', cleaned):
        return True
    # "황소영/정종현" 같은 복수 강사
    if re.match(r'^[가-힣]{2,4}/[가-힣]{2,4}$', cleaned):
        return True
    # "박정일강사" 같은 패턴 → "박정일"로 정제
    if re.match(r'^[가-힣]{2,4}강사$', cleaned):
        return True
    return False


def _extract_names_from_text(text):
    """텍스트에서 한국어 강사명 추출 (복수 가능)"""
    names = []

    # "황소영/정종현" 슬래시 구분
    if '/' in text:
        for part in text.split('/'):
            part = part.strip()
            if part.endswith('강사'):
                part = part[:-2]
            if re.match(r'^[가-힣]{2,4}$', part):
                names.append(part)
        if names:
            return names

    # "강명호,인선미" 쉼표 구분
    if ',' in text:
        for part in text.split(','):
            part = part.strip().split()[0] if part.strip() else ''
            if part.endswith('강사'):
                part = part[:-2]
            if re.match(r'^[가-힣]{2,4}$', part):
                names.append(part)
        if names:
            return names

    # 단어별 스캔
    for part in text.split():
        cleaned = part.strip()
        if cleaned.endswith('강사'):
            cleaned = cleaned[:-2]
        if cleaned in NOT_NAME_WORDS:
            continue
        if re.match(r'^[가-힣]{2,4}$', cleaned):
            names.append(cleaned)

    return names


def _extract_instructors_and_hours(ws, date_col, class_col, week_row, default_hours):
    """수업명 아래에서 강사명(복수)과 수업시간 추출"""
    instructors = []
    numeric_hours = None      # 숫자 셀에서 발견된 시간 (우선순위 1)
    text_hours_list = []      # 텍스트 "Xh"에서 발견된 시간들

    # 날짜 열과 수업명 열 모두에서 스캔
    scan_cols = sorted(set([date_col, class_col]))

    for offset in range(1, 6):
        row = week_row + offset
        for col in scan_cols:
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is None:
                continue

            if isinstance(cell_val, (int, float)):
                # 숫자: 수업시간 (1~12 범위)
                val = int(cell_val)
                if 1 <= val <= 12 and numeric_hours is None:
                    numeric_hours = val
            elif isinstance(cell_val, str):
                text = cell_val.strip()
                if not text:
                    continue

                # 강사명 추출 (복수 수집)
                names = _extract_names_from_text(text)
                for name in names:
                    if name not in instructors:
                        instructors.append(name)

                # 텍스트 내 시간 추출 (12h 이하만 — 16h 버그 수정)
                for part in text.split():
                    m = re.match(r'^(\d+)h$', part, re.IGNORECASE)
                    if m:
                        h = int(m.group(1))
                        if 1 <= h <= 12:
                            text_hours_list.append(h)

    # 시간 결정: 숫자 셀 우선 → 텍스트 합산 → 기본값
    if numeric_hours is not None:
        hours = numeric_hours
    elif text_hours_list:
        hours = sum(text_hours_list)
    else:
        hours = default_hours

    # 강사 있고 시간 없으면 기본값
    if instructors and numeric_hours is None and not text_hours_list:
        logger.debug(f"  강사 {instructors} 발견, 시간 미지정 → 기본 {hours}h 적용")

    instructor_str = ",".join(instructors)
    return instructor_str, hours


def parse_timetable(filepath, selected_sheets, default_start_time='09:00', default_hours=8):
    """엑셀 시간표 파싱 → ClassEntry 리스트 반환"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    entries = []

    for sheet_name in selected_sheets:
        if sheet_name in INFO_SHEETS:
            continue
        if sheet_name not in wb.sheetnames:
            logger.warning(f"시트 '{sheet_name}'을 찾을 수 없습니다.")
            continue

        ws = wb[sheet_name]
        logger.info(f"시트 '{sheet_name}' 파싱 중...")

        # 동적으로 주 시작 행 감지
        week_start_rows = _detect_week_start_rows(ws)

        for week_row in week_start_rows:
            for date_col, class_col in DAY_CONFIG:
                # 1. 날짜 읽기
                date_cell = ws.cell(row=week_row, column=date_col).value
                if date_cell is None:
                    continue
                if not isinstance(date_cell, datetime):
                    continue
                if date_cell.year < 2020:
                    continue

                # 2. 수업명 읽기
                class_cell = ws.cell(row=week_row, column=class_col).value
                if class_cell is None or not isinstance(class_cell, str):
                    continue
                class_name = class_cell.strip().replace('\n', ' ').replace('\t', ' ')
                class_name = re.sub(r'\s+', ' ', class_name)
                if not class_name:
                    continue

                # 3. 공휴일 체크
                is_holiday = any(kw in class_name for kw in HOLIDAY_KEYWORDS)

                # 4. 수업명에서 잔여 시간 패턴 제거 ("수업명 3h" → "수업명")
                class_name = re.sub(r'\s+\d+h\s*$', '', class_name, flags=re.IGNORECASE)

                # 5. 강사명(복수), 수업시간 추출
                instructor = ""
                hours = default_hours
                if not is_holiday:
                    instructor, hours = _extract_instructors_and_hours(
                        ws, date_col, class_col, week_row, default_hours
                    )

                # 6. 종료 시간 계산
                end_time = calculate_end_time(default_start_time, hours)

                # 7. Entry 생성
                entry = {
                    "date": date_cell.strftime('%Y-%m-%d'),
                    "class_name": class_name,
                    "instructor": instructor,
                    "hours": hours,
                    "start_time": default_start_time,
                    "end_time": end_time,
                    "is_holiday": is_holiday,
                }
                entries.append(entry)
                logger.debug(f"  {entry['date']} | {class_name} | {instructor} | {hours}h")

    wb.close()
    logger.info(f"총 {len(entries)}개 수업 일정 파싱 완료")
    return entries
